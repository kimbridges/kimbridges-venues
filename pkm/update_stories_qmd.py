"""
Update .qmd story files from stories_inventory_v2.xlsx

Reads Card Subtitle and Full Description columns.
- Card Subtitle → YAML subtitle: field (truncated to ~90 chars with "...")
- Full Description → body text above the page count and viewer

Usage:
    python update_stories_qmd.py

Expects:
    - stories_inventory_v2.xlsx in the same directory as this script
    - ../gePoints/kimbridges-stories/stories/ contains the .qmd files
"""

import os
import re
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(SCRIPT_DIR, 'stories_inventory_v2.xlsx')
STORIES_DIR = os.path.join(SCRIPT_DIR, '..', 'gePoints', 'kimbridges-stories', 'stories')

MAX_SUBTITLE = 90

def truncate(text, max_len):
    if not text or len(text) <= max_len:
        return text
    return text[:max_len].rsplit(' ', 1)[0] + '...'

def read_spreadsheet():
    wb = load_workbook(XLSX)
    ws = wb.active
    stories = {}
    for r in range(2, ws.max_row + 1):
        title = ws.cell(row=r, column=3).value
        card = ws.cell(row=r, column=6).value
        full = ws.cell(row=r, column=7).value
        pages = ws.cell(row=r, column=4).value
        if title:
            stories[title] = {
                'card_subtitle': card or '',
                'full_description': full or card or '',
                'pages': pages
            }
    return stories

def get_qmd_title(filepath):
    with open(filepath) as f:
        content = f.read()
    m = re.search(r'^title:\s*"(.+?)"', content, re.MULTILINE)
    return m.group(1) if m else None

def update_qmd(filepath, story_data):
    with open(filepath) as f:
        content = f.read()

    # Split into YAML front matter and body
    parts = content.split('---', 2)
    if len(parts) < 3:
        return False

    yaml_block = parts[1]
    body = parts[2]

    # Update subtitle in YAML
    short = truncate(story_data['card_subtitle'], MAX_SUBTITLE)
    short_escaped = short.replace('"', '\\"')
    yaml_block = re.sub(
        r'^subtitle:\s*".*?"',
        f'subtitle: "{short_escaped}"',
        yaml_block,
        flags=re.MULTILINE
    )

    # Update body: replace the description text before "Pages:"
    full_desc = story_data['full_description']
    body = re.sub(
        r'^\n.*?\n\nPages:',
        f'\n{full_desc}\n\nPages:',
        body,
        count=1,
        flags=re.DOTALL
    )

    new_content = '---' + yaml_block + '---' + body
    with open(filepath, 'w') as f:
        f.write(new_content)
    return True

def main():
    stories = read_spreadsheet()
    print(f"Read {len(stories)} stories from spreadsheet")

    qmd_files = [f for f in os.listdir(STORIES_DIR) if f.endswith('.qmd')]
    updated = 0
    skipped = []

    for qf in sorted(qmd_files):
        path = os.path.join(STORIES_DIR, qf)
        title = get_qmd_title(path)
        if title and title in stories:
            if update_qmd(path, stories[title]):
                updated += 1
                print(f"  Updated: {qf}")
            else:
                skipped.append(qf)
                print(f"  Skipped (parse error): {qf}")
        else:
            skipped.append(qf)
            print(f"  Skipped (no match): {qf} → {title}")

    print(f"\nDone. Updated {updated}, skipped {len(skipped)}")

if __name__ == '__main__':
    main()
