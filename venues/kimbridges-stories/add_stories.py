"""
Add new stories to the kimbridges-stories Quarto site.

For each PDF that lacks a .qmd file:
  1. Generate a thumbnail (first page → JPG) in images/
  2. Create a .qmd file in stories/ using spreadsheet metadata
  3. If the PDF is in updates/, move it to pdfs/

Usage:
    python add_stories.py [--dry-run]

Expects:
    - stories_inventory_v2.xlsx in the project root
    - pdfs/    — published PDFs
    - updates/ — new PDFs to be added
    - images/  — thumbnail JPGs
    - stories/ — .qmd story pages
"""

import os
import re
import sys
import shutil
import fitz  # pymupdf
from openpyxl import load_workbook

# --- Configuration ---
PROJECT_DIR = os.environ.get('PROJECT_DIR_OVERRIDE',
                             os.path.dirname(os.path.abspath(__file__)))
XLSX = os.environ.get('XLSX_OVERRIDE',
                      os.path.join(PROJECT_DIR, 'stories_inventory_v2.xlsx'))
PDFS_DIR = os.path.join(PROJECT_DIR, 'pdfs')
UPDATES_DIR = os.path.join(PROJECT_DIR, 'updates')
IMAGES_DIR = os.path.join(PROJECT_DIR, 'images')
STORIES_DIR = os.path.join(PROJECT_DIR, 'stories')

MAX_SUBTITLE = 90
THUMB_DPI = 150


def truncate(text, max_len):
    """Truncate text to max_len, breaking at word boundary."""
    if not text or len(text) <= max_len:
        return text
    return text[:max_len].rsplit(' ', 1)[0] + '...'


def slugify(title):
    """Convert a title to a URL-friendly slug for the .qmd filename."""
    slug = title.lower()
    slug = slug.replace("'", '')      # Cruis'n → Cruisn
    slug = slug.replace('`', '')      # Pa`u → Pau
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    slug = slug.strip('-')
    return slug


def read_spreadsheet():
    """Read story metadata from the spreadsheet, keyed by PDF filename."""
    wb = load_workbook(XLSX)
    ws = wb.active
    stories = {}
    for r in range(2, ws.max_row + 1):
        title = ws.cell(row=r, column=3).value
        pages = ws.cell(row=r, column=4).value
        category = ws.cell(row=r, column=5).value
        card_subtitle = ws.cell(row=r, column=6).value
        full_desc = ws.cell(row=r, column=7).value
        pdf_name = ws.cell(row=r, column=8).value
        if title and pdf_name:
            # Normalize PDF filename
            if not pdf_name.endswith('.pdf'):
                pdf_name += '.pdf'
            stories[pdf_name] = {
                'title': title,
                'pages': pages,
                'categories': parse_categories(category),
                'card_subtitle': card_subtitle or '',
                'full_description': full_desc or card_subtitle or '',
            }
    return stories


def parse_categories(cat_str):
    """Parse 'Travel / Food / Japan' into ['Travel', 'Food', 'Japan']."""
    if not cat_str:
        return []
    return [c.strip() for c in cat_str.split('/') if c.strip()]


def generate_thumbnail(pdf_path, image_path):
    """Render first page of PDF as a JPG thumbnail."""
    doc = fitz.open(pdf_path)
    page = doc[0]
    mat = fitz.Matrix(THUMB_DPI / 72, THUMB_DPI / 72)
    pix = page.get_pixmap(matrix=mat)
    pix.save(image_path)
    doc.close()


def create_qmd(slug, story_data, pdf_filename):
    """Create a .qmd file for a story."""
    title = story_data['title']
    subtitle = truncate(story_data['card_subtitle'], MAX_SUBTITLE)
    categories = story_data['categories']
    full_desc = story_data['full_description']
    pages = story_data['pages'] or ''

    # Determine image filename (same stem as PDF, .jpg)
    image_stem = os.path.splitext(pdf_filename)[0]
    image_file = f'{image_stem}.jpg'

    # Escape quotes in YAML strings
    title_esc = title.replace('"', '\\"')
    subtitle_esc = subtitle.replace('"', '\\"')
    cat_yaml = ', '.join(f'"{c}"' for c in categories)

    # Build date from today or leave blank
    from datetime import date
    today = date.today().isoformat()

    content = f'''---
title: "{title_esc}"
subtitle: "{subtitle_esc}"
date: {today}
categories: [{cat_yaml}]
image: ../images/{image_file}
pdf: ../pdfs/{pdf_filename}
---
{full_desc}

Pages: {pages}

<iframe src="../viewer.html?pdf={pdf_filename}" width="100%" height="700px" style="border: none; border-radius: 6px;"></iframe>
'''

    qmd_path = os.path.join(STORIES_DIR, f'{slug}.qmd')
    with open(qmd_path, 'w') as f:
        f.write(content)
    return qmd_path


def get_existing_qmd_pdfs():
    """Return set of PDF filenames referenced by existing .qmd files."""
    refs = set()
    for f in os.listdir(STORIES_DIR):
        if f.endswith('.qmd'):
            with open(os.path.join(STORIES_DIR, f)) as fh:
                m = re.search(r'^pdf:\s*\.\./pdfs/(.+\.pdf)', fh.read(), re.MULTILINE)
                if m:
                    refs.add(m.group(1))
    return refs


def main():
    dry_run = '--dry-run' in sys.argv

    # Read spreadsheet
    stories = read_spreadsheet()
    print(f'Spreadsheet: {len(stories)} entries with PDF filenames')

    # Find existing .qmd references
    existing_refs = get_existing_qmd_pdfs()
    print(f'Existing .qmd files referencing: {len(existing_refs)} PDFs')

    # Collect all PDFs needing processing
    to_process = []

    # PDFs in updates/
    if os.path.isdir(UPDATES_DIR):
        for f in sorted(os.listdir(UPDATES_DIR)):
            if f.endswith('.pdf') and f not in existing_refs:
                to_process.append(('updates', f))

    # PDFs in pdfs/ without .qmd
    for f in sorted(os.listdir(PDFS_DIR)):
        if f.endswith('.pdf') and f not in existing_refs:
            to_process.append(('pdfs', f))

    if not to_process:
        print('\nNothing to process — all PDFs have .qmd files.')
        return

    print(f'\nProcessing {len(to_process)} PDFs:\n')

    added = 0
    skipped = []

    for source, pdf_file in to_process:
        # Look up in spreadsheet
        if pdf_file not in stories:
            skipped.append((pdf_file, 'not in spreadsheet'))
            print(f'  SKIP: {pdf_file} (not in spreadsheet)')
            continue

        data = stories[pdf_file]
        slug = slugify(data['title'])
        qmd_path = os.path.join(STORIES_DIR, f'{slug}.qmd')
        pdf_stem = os.path.splitext(pdf_file)[0]
        image_file = f'{pdf_stem}.jpg'
        image_path = os.path.join(IMAGES_DIR, image_file)

        if dry_run:
            print(f'  [DRY RUN] {pdf_file}')
            print(f'           slug: {slug}')
            print(f'           title: {data["title"]}')
            print(f'           source: {source}/')
            if source == 'updates':
                print(f'           → move to pdfs/')
            print(f'           → generate thumbnail: {image_file}')
            print(f'           → create: stories/{slug}.qmd')
            print()
            added += 1
            continue

        # Determine PDF location
        if source == 'updates':
            pdf_path = os.path.join(UPDATES_DIR, pdf_file)
        else:
            pdf_path = os.path.join(PDFS_DIR, pdf_file)

        # 1. Generate thumbnail if missing
        if not os.path.exists(image_path):
            try:
                generate_thumbnail(pdf_path, image_path)
                print(f'  Thumbnail: {image_file}')
            except Exception as e:
                skipped.append((pdf_file, f'thumbnail error: {e}'))
                print(f'  SKIP: {pdf_file} (thumbnail error: {e})')
                continue
        else:
            print(f'  Thumbnail exists: {image_file}')

        # 2. Create .qmd
        if not os.path.exists(qmd_path):
            create_qmd(slug, data, pdf_file)
            print(f'  Created: stories/{slug}.qmd')
        else:
            print(f'  .qmd exists: {slug}.qmd')

        # 3. Move PDF from updates/ to pdfs/
        if source == 'updates':
            dest = os.path.join(PDFS_DIR, pdf_file)
            if not os.path.exists(dest):
                shutil.move(pdf_path, dest)
                print(f'  Moved: updates/{pdf_file} → pdfs/{pdf_file}')
            else:
                print(f'  Already in pdfs/: {pdf_file}')

        added += 1
        print()

    print(f'Done. Processed {added}, skipped {len(skipped)}')
    if skipped:
        print('Skipped:')
        for name, reason in skipped:
            print(f'  {name}: {reason}')


if __name__ == '__main__':
    main()
